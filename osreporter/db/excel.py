"""
Write data to an Excel workbook.
"""
from datetime import datetime
import os

from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill


def usage(data):
    """Write data to an Excel spreadsheet."""
    highlight = NamedStyle(name='highlight')
    highlight.font = Font(bold=True, size=20)

    tabletop = NamedStyle(name='tabletop')
    tabletop.font = Font(bold=True, color='FFFFFF')
    bd = Side(style='thin', color="000000", border_style='thin')
    tabletop.border = Border(left=bd, right=bd, top=bd, bottom=bd)
    tabletop.fill = PatternFill("solid", bgColor="333333")

    wb = Workbook()
    wb.add_named_style(highlight)

    current_row = 4

    for sheet in wb:
        if "Sheet" in sheet.title:
            wb.remove_sheet(sheet)

    ws = wb.create_sheet(title="Usage", index=0)
    ws.sheet_properties.tabColor = "0000FF"

    ws['A1'].style = 'highlight'
    ws['A1'] = "CVM Usage"

    ws['A2'] = ""

    headings = ['Name', 'Type', 'Instances', 'vCPU', 'RAM', 'Local Storage', 'Shared Storage', 'Active', 'Suspended', 'Shutdown', 'Error', 'Other']
    ws.append(headings)
    header_row = ws[3]
    for cell in header_row:
        cell.style = tabletop

    # Set column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 11
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 11
    ws.column_dimensions["F"].width = 11
    ws.column_dimensions["G"].width = 11
    ws.column_dimensions["H"].width = 11
    ws.column_dimensions["I"].width = 11
    ws.column_dimensions["J"].width = 11
    ws.column_dimensions["K"].width = 11
    ws.column_dimensions["L"].width = 11

    for point in data:
        ws["A{0}".format(current_row)] = point[0]
        ws["A{0}".format(current_row)].style = 'Output'
        ws["B{0}".format(current_row)] = point[1]
        if 'Internal' in point[1]:
            ws["B{0}".format(current_row)].style = 'Accent1'
        elif 'External' in point[1]:
            ws["B{0}".format(current_row)].style = 'Accent3'
        else:
            ws["B{0}".format(current_row)].style = 'Accent5'
        ws["C{0}".format(current_row)] = point[3]
        ws["D{0}".format(current_row)] = point[4]
        ws["E{0}".format(current_row)] = point[5]
        ws["F{0}".format(current_row)] = point[6]
        ws["G{0}".format(current_row)] = point[7]
        ws["H{0}".format(current_row)] = point[8]
        ws["I{0}".format(current_row)] = point[9]
        if point[9] > 0:
            ws["I{0}".format(current_row)].style = 'Neutral'
        else:
            ws["I{0}".format(current_row)].style = 'Good'
        ws["J{0}".format(current_row)] = point[10]
        if point[10] > 0:
            ws["J{0}".format(current_row)].style = 'Bad'
        else:
            ws["J{0}".format(current_row)].style = 'Good'
        ws["K{0}".format(current_row)] = point[11]
        if point[11] > 0:
            ws["K{0}".format(current_row)].style = 'Bad'
        else:
            ws["K{0}".format(current_row)].style = 'Good'
        ws["L{0}".format(current_row)] = point[12]
        if point[12] > 0:
            ws["L{0}".format(current_row)].style = 'Neutral'
        else:
            ws["L{0}".format(current_row)].style = 'Good'
        current_row += 1

    filename = os.path.join(os.path.expanduser('~'), "CVM Usage Report {0}.xlsx".format(datetime.now().strftime("%Y-%m-%d")))
    wb.save(filename)


def flavors(data):
    """Write flavor data to an Excel spreadsheet. """

    highlight = NamedStyle(name='highlight')
    highlight.font = Font(bold=True, size=20)

    tabletop = NamedStyle(name='tabletop')
    tabletop.font = Font(bold=True, color='FFFFFF')
    bd = Side(style='thin', color="000000", border_style='thin')
    tabletop.border = Border(left=bd, right=bd, top=bd, bottom=bd)
    tabletop.fill = PatternFill("solid", bgColor="333333")

    wb = Workbook()
    wb.add_named_style(highlight)

    current_row = 4

    for sheet in wb:
        if "Sheet" in sheet.title:
            wb.remove_sheet(sheet)

    ws = wb.create_sheet(title="Usage", index=0)
    ws.sheet_properties.tabColor = "0000FF"

    ws['A1'].style = 'highlight'
    ws['A1'] = "CVM Flavor Distribution"

    ws['A2'] = ""

    headings = ['Name', 'Total']
    ws.append(headings)
    header_row = ws[3]
    for cell in header_row:
        cell.style = tabletop

    # Set column widths
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 10

    for key, value in data.items():
        ws["A{0}".format(current_row)] = key
        ws["A{0}".format(current_row)].style = 'Output'
        ws["B{0}".format(current_row)] = value
        if value > 20:
            ws["B{0}".format(current_row)].style = 'Good'
        else:
            ws["B{0}".format(current_row)].style = 'Neutral'
        current_row += 1

    filename = os.path.join(os.path.expanduser('~'), "CVM Flavor Distribition Report {0}.xlsx".format(datetime.now().strftime("%Y-%m-%d")))
    wb.save(filename)
